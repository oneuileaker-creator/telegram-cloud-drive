# app/services/ftms/metadata.py

import io
import logging
from typing import Optional
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)


@dataclass
class ImageMetadata:
    width: int = 0
    height: int = 0
    mode: str = ""
    format: str = ""
    has_transparency: bool = False
    exif: dict = field(default_factory=dict)
    gps: Optional[dict] = None
    camera: Optional[str] = None
    taken_at: Optional[str] = None
    iso: Optional[int] = None
    aperture: Optional[str] = None
    shutter_speed: Optional[str] = None


@dataclass
class VideoMetadata:
    duration_seconds: float = 0.0
    width: int = 0
    height: int = 0
    fps: float = 0.0
    bitrate: int = 0
    video_codec: str = ""
    audio_codec: str = ""
    has_audio: bool = False


@dataclass
class AudioMetadata:
    duration_seconds: float = 0.0
    bitrate: int = 0
    sample_rate: int = 0
    channels: int = 0
    codec: str = ""
    title: Optional[str] = None
    artist: Optional[str] = None
    album: Optional[str] = None
    year: Optional[str] = None
    genre: Optional[str] = None
    track_number: Optional[str] = None
    has_cover_art: bool = False


@dataclass
class DocumentMetadata:
    page_count: int = 0
    word_count: int = 0
    title: Optional[str] = None
    author: Optional[str] = None
    created_at: Optional[str] = None
    modified_at: Optional[str] = None
    has_images: bool = False


@dataclass
class ArchiveMetadata:
    file_count: int = 0
    total_uncompressed_bytes: int = 0
    compression_ratio: float = 0.0
    file_list: list = field(default_factory=list)


class MetadataExtractor:

    # ─── Main Entry ───────────────────────────────────────────

    @staticmethod
    async def extract(
        file_data: bytes,
        category: str,
        mime_type: str
    ) -> dict:
        """
        Extract metadata from file based on category.
        Always returns a dict (empty if unsupported).
        """
        try:
            if category == "image":
                return await MetadataExtractor._image(file_data)
            elif category == "video":
                return await MetadataExtractor._video(file_data)
            elif category == "audio":
                return await MetadataExtractor._audio(file_data)
            elif category == "document":
                return await MetadataExtractor._document(file_data, mime_type)
            elif category == "archive":
                return await MetadataExtractor._archive(file_data, mime_type)
            elif category == "code":
                return await MetadataExtractor._code(file_data)
            return {}
        except Exception as e:
            logger.warning(f"Metadata extraction failed ({category}): {e}")
            return {}

    # ─── Image Metadata ───────────────────────────────────────

    @staticmethod
    async def _image(file_data: bytes) -> dict:
        import asyncio

        def _process():
            from PIL import Image
            from PIL.ExifTags import TAGS, GPSTAGS

            result = ImageMetadata()

            try:
                img = Image.open(io.BytesIO(file_data))
                result.width = img.width
                result.height = img.height
                result.mode = img.mode
                result.format = img.format or ""
                result.has_transparency = img.mode in ("RGBA", "LA", "PA")

                # Extract EXIF
                exif_data = img._getexif()
                if exif_data:
                    for tag_id, value in exif_data.items():
                        tag = TAGS.get(tag_id, tag_id)

                        # GPS data
                        if tag == "GPSInfo":
                            gps = {}
                            for gps_id, gps_val in value.items():
                                gps[GPSTAGS.get(gps_id, gps_id)] = str(gps_val)
                            result.gps = gps
                            continue

                        # Camera info
                        if tag in ("Make", "Model"):
                            result.camera = (result.camera or "") + str(value) + " "

                        # Shot info
                        if tag == "DateTimeOriginal":
                            result.taken_at = str(value)
                        if tag == "ISOSpeedRatings":
                            result.iso = int(value)
                        if tag == "FNumber":
                            result.aperture = f"f/{float(value):.1f}"
                        if tag == "ExposureTime":
                            result.shutter_speed = f"1/{int(1/float(value))}s" if value < 1 else f"{float(value)}s"

                        # Store clean EXIF (skip binary)
                        if isinstance(value, (str, int, float)):
                            result.exif[tag] = value

            except Exception as e:
                logger.debug(f"Image metadata partial error: {e}")

            return result.__dict__

        loop = __import__("asyncio").get_event_loop()
        return await loop.run_in_executor(None, _process)

    # ─── Video Metadata ───────────────────────────────────────

    @staticmethod
    async def _video(file_data: bytes) -> dict:
        import asyncio
        import tempfile
        import os
        import json

        result = VideoMetadata()

        with tempfile.NamedTemporaryFile(
            suffix=".mp4",
            delete=False
        ) as tmp:
            tmp.write(file_data)
            tmp_path = tmp.name

        try:
            proc = await asyncio.create_subprocess_exec(
                "ffprobe",
                "-v", "quiet",
                "-print_format", "json",
                "-show_streams",
                "-show_format",
                tmp_path,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE
            )
            stdout, _ = await proc.communicate()
            probe = json.loads(stdout)

            fmt = probe.get("format", {})
            result.duration_seconds = float(fmt.get("duration", 0))
            result.bitrate = int(fmt.get("bit_rate", 0))

            for stream in probe.get("streams", []):
                if stream.get("codec_type") == "video":
                    result.width = stream.get("width", 0)
                    result.height = stream.get("height", 0)
                    result.video_codec = stream.get("codec_name", "")
                    fps_str = stream.get("r_frame_rate", "0/1")
                    try:
                        num, den = fps_str.split("/")
                        result.fps = round(int(num) / int(den), 2)
                    except Exception:
                        result.fps = 0.0

                elif stream.get("codec_type") == "audio":
                    result.audio_codec = stream.get("codec_name", "")
                    result.has_audio = True

        except Exception as e:
            logger.warning(f"Video metadata error: {e}")
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

        return result.__dict__

    # ─── Audio Metadata ───────────────────────────────────────

    @staticmethod
    async def _audio(file_data: bytes) -> dict:
        import asyncio

        def _process():
            from mutagen import File as MutagenFile

            result = AudioMetadata()
            audio = MutagenFile(io.BytesIO(file_data))

            if not audio:
                return result.__dict__

            # Duration & technical info
            if audio.info:
                result.duration_seconds = round(audio.info.length, 2)
                result.bitrate = getattr(audio.info, "bitrate", 0)
                result.sample_rate = getattr(audio.info, "sample_rate", 0)
                result.channels = getattr(audio.info, "channels", 0)

            # Tags
            tags = audio.tags
            if tags:
                def get_tag(*keys):
                    for key in keys:
                        val = tags.get(key)
                        if val:
                            return str(val[0]) if isinstance(val, list) else str(val)
                    return None

                result.title = get_tag("TIT2", "title", "\xa9nam")
                result.artist = get_tag("TPE1", "artist", "\xa9ART")
                result.album = get_tag("TALB", "album", "\xa9alb")
                result.year = get_tag("TDRC", "date", "\xa9day")
                result.genre = get_tag("TCON", "genre", "\xa9gen")
                result.track_number = get_tag("TRCK", "tracknumber", "trkn")

                # Check for cover art
                for key in tags.keys():
                    if key.startswith("APIC") or key == "covr":
                        result.has_cover_art = True
                        break

            return result.__dict__

        loop = __import__("asyncio").get_event_loop()
        return await loop.run_in_executor(None, _process)

    # ─── Document Metadata ────────────────────────────────────

    @staticmethod
    async def _document(file_data: bytes, mime_type: str) -> dict:
        import asyncio

        result = DocumentMetadata()

        def _process_pdf():
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(file_data))
            result.page_count = len(reader.pages)

            info = reader.metadata
            if info:
                result.title = info.get("/Title")
                result.author = info.get("/Author")
                result.created_at = str(info.get("/CreationDate", ""))
                result.modified_at = str(info.get("/ModDate", ""))

            # Count words from first 5 pages
            text = ""
            for page in reader.pages[:5]:
                text += page.extract_text() or ""
            result.word_count = len(text.split())
            return result.__dict__

        def _process_docx():
            from docx import Document
            doc = Document(io.BytesIO(file_data))
            result.page_count = 1         # python-docx doesn't give page count
            result.title = doc.core_properties.title
            result.author = doc.core_properties.author
            result.created_at = str(doc.core_properties.created or "")
            result.modified_at = str(doc.core_properties.modified or "")
            text = "\n".join([p.text for p in doc.paragraphs])
            result.word_count = len(text.split())
            result.has_images = len(doc.inline_shapes) > 0
            return result.__dict__

        def _process_xlsx():
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
            result.page_count = len(wb.sheetnames)
            return result.__dict__

        loop = __import__("asyncio").get_event_loop()

        try:
            if mime_type == "application/pdf":
                return await loop.run_in_executor(None, _process_pdf)
            elif "wordprocessing" in mime_type or mime_type == "application/msword":
                return await loop.run_in_executor(None, _process_docx)
            elif "spreadsheet" in mime_type or "excel" in mime_type:
                return await loop.run_in_executor(None, _process_xlsx)
        except Exception as e:
            logger.warning(f"Document metadata error: {e}")

        return result.__dict__

    # ─── Archive Metadata ─────────────────────────────────────

    @staticmethod
    async def _archive(file_data: bytes, mime_type: str) -> dict:
        import asyncio

        def _process():
            result = ArchiveMetadata()

            try:
                if "zip" in mime_type:
                    import zipfile
                    with zipfile.ZipFile(io.BytesIO(file_data)) as zf:
                        infos = zf.infolist()
                        result.file_count = len(infos)
                        result.total_uncompressed_bytes = sum(
                            i.file_size for i in infos
                        )
                        result.file_list = [i.filename for i in infos[:100]]
                        if result.total_uncompressed_bytes > 0:
                            result.compression_ratio = round(
                                len(file_data) / result.total_uncompressed_bytes, 2
                            )

                elif "x-tar" in mime_type or "gzip" in mime_type:
                    import tarfile
                    with tarfile.open(fileobj=io.BytesIO(file_data)) as tf:
                        members = tf.getmembers()
                        result.file_count = len(members)
                        result.file_list = [m.name for m in members[:100]]

            except Exception as e:
                logger.debug(f"Archive read error: {e}")

            return result.__dict__

        loop = __import__("asyncio").get_event_loop()
        return await loop.run_in_executor(None, _process)

    # ─── Code Metadata ────────────────────────────────────────

    @staticmethod
    async def _code(file_data: bytes) -> dict:
        try:
            text = file_data.decode("utf-8", errors="replace")
            lines = text.split("\n")
            return {
                "line_count": len(lines),
                "char_count": len(text),
                "word_count": len(text.split()),
                "blank_lines": sum(1 for l in lines if not l.strip()),
                "preview": "\n".join(lines[:10])
            }
        except Exception:
            return {}
